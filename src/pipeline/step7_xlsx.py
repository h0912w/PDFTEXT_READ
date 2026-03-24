"""
Step 7: Export review_output.xlsx.

Layout rules (from io-contracts.md):
  LEFT_TO_RIGHT pages → top image / bottom text
  TOP_TO_BOTTOM pages → left image / right text

One worksheet per page.
Images are embedded; text blocks appear adjacent with order_index,
text, confidence, status, source.
Skipped/Unknown blocks are highlighted in orange for easy review.

State transition: VALIDATED|APPROVED_WITH_WARNINGS → EXPORT_COMPLETED
"""
from __future__ import annotations

import os
from typing import List

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.models.state import (
    PipelineContext,
    ProcessingStatus,
    ReadingDirection,
    TextBlock,
    TextStatus,
)
from src.utils.image_utils import resize_image_for_xlsx
from src.utils.logger import get_logger

_XLSX_FILENAME = "review_output.xlsx"

# Image display size in pixels (for embedding)
_IMG_MAX_W = 540
_IMG_MAX_H = 720

# Excel row height in points and column width in chars
_ROW_HEIGHT = 15.0
_IMG_COL_WIDTH = 8.0   # pixels-per-char approximation: ~7px

# Color fills
_FILL_HEADER = PatternFill("solid", fgColor="4472C4")
_FILL_SKIP = PatternFill("solid", fgColor="FFD966")
_FILL_UNKNOWN = PatternFill("solid", fgColor="FF7070")
_FILL_REVIEW = PatternFill("solid", fgColor="FFF2CC")

_FONT_HEADER = Font(bold=True, color="FFFFFF")
_FONT_BOLD = Font(bold=True)


def run(ctx: PipelineContext) -> PipelineContext:
    """Write review_output.xlsx to the work directory."""
    logger = get_logger()
    logger.info("Step 7: Generating review_output.xlsx…")

    xlsx_path = os.path.join(ctx.work_dir, _XLSX_FILENAME)

    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default empty sheet

        # Group blocks by page
        blocks_by_page: dict = {}
        for tb in sorted(ctx.text_blocks, key=lambda b: b.order_index):
            blocks_by_page.setdefault(tb.page_num, []).append(tb)

        for pi in ctx.page_infos:
            page_blocks = blocks_by_page.get(pi.page_num, [])
            ws = wb.create_sheet(title=f"Page_{pi.page_num:03d}")

            img_path = pi.image_path
            if img_path and os.path.exists(img_path):
                thumb_path, img_w, img_h = resize_image_for_xlsx(img_path, _IMG_MAX_W, _IMG_MAX_H)
            else:
                thumb_path = None
                img_w, img_h = _IMG_MAX_W, _IMG_MAX_H

            if pi.direction == ReadingDirection.LEFT_TO_RIGHT.value:
                _layout_ltr(ws, thumb_path, img_w, img_h, page_blocks, pi)
            else:
                _layout_ttb(ws, thumb_path, img_w, img_h, page_blocks, pi)

        wb.save(xlsx_path)
        logger.info(f"  XLSX written: {xlsx_path}")

        ctx.options["_xlsx_path"] = xlsx_path
        ctx.status = ProcessingStatus.EXPORT_COMPLETED
        logger.info(f"Step 7 complete. Status: {ctx.status}")

    except Exception as exc:
        ctx.status = ProcessingStatus.FAILED
        ctx.add_error(f"Step 7 failed: {exc}")
        logger.error(f"Step 7 failed: {exc}", exc_info=True)

    return ctx


# ── Layout builders ───────────────────────────────────────────────────────

def _layout_ltr(ws, thumb_path, img_w: int, img_h: int, blocks: List[TextBlock], pi) -> None:
    """
    LEFT_TO_RIGHT: image on top, text table below.
    """
    # Calculate rows the image will occupy
    px_per_row = _ROW_HEIGHT * 1.33  # ~1.33 px per point at 96 DPI
    img_rows = max(1, int(img_h / px_per_row))
    img_cols = max(1, int(img_w / (_IMG_COL_WIDTH * 7)))  # ~7 px per col-unit

    # Set column widths
    for col in range(1, img_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = _IMG_COL_WIDTH
    # Text columns after image span
    text_start_col = 1
    for i, col_width in enumerate([8, 40, 12, 12, 14]):
        ws.column_dimensions[get_column_letter(text_start_col + i)].width = col_width

    # Set row heights for image area
    for r in range(1, img_rows + 2):
        ws.row_dimensions[r].height = _ROW_HEIGHT

    # Embed image
    if thumb_path and os.path.exists(thumb_path):
        xl_img = XLImage(thumb_path)
        xl_img.width = img_w
        xl_img.height = img_h
        xl_img.anchor = "A1"
        ws.add_image(xl_img)

    # Merge info label
    header_row = img_rows + 2
    ws.cell(row=header_row - 1, column=1, value=f"Page {pi.page_num} | {pi.doc_type} | {pi.direction}")
    ws.cell(row=header_row - 1, column=1).font = _FONT_BOLD

    # Text table header
    _write_text_header(ws, header_row, start_col=1)

    # Text rows
    for i, tb in enumerate(blocks):
        row = header_row + 1 + i
        _write_text_row(ws, row, start_col=1, tb=tb)


def _layout_ttb(ws, thumb_path, img_w: int, img_h: int, blocks: List[TextBlock], pi) -> None:
    """
    TOP_TO_BOTTOM: image on the left, text table on the right.
    """
    px_per_col_unit = _IMG_COL_WIDTH * 7
    img_col_count = max(1, int(img_w / px_per_col_unit))
    px_per_row = _ROW_HEIGHT * 1.33
    img_row_count = max(1, int(img_h / px_per_row))

    # Set image column widths
    for col in range(1, img_col_count + 1):
        ws.column_dimensions[get_column_letter(col)].width = _IMG_COL_WIDTH

    # Set row heights
    total_rows = max(img_row_count, len(blocks) + 5)
    for r in range(1, total_rows + 2):
        ws.row_dimensions[r].height = _ROW_HEIGHT

    # Embed image
    if thumb_path and os.path.exists(thumb_path):
        xl_img = XLImage(thumb_path)
        xl_img.width = img_w
        xl_img.height = img_h
        xl_img.anchor = "A1"
        ws.add_image(xl_img)

    # Text table starts right of image
    text_col_start = img_col_count + 2
    for i, col_width in enumerate([8, 40, 12, 12, 14]):
        ws.column_dimensions[get_column_letter(text_col_start + i)].width = col_width

    # Info label
    ws.cell(row=1, column=text_col_start, value=f"Page {pi.page_num} | {pi.doc_type} | {pi.direction}")
    ws.cell(row=1, column=text_col_start).font = _FONT_BOLD

    # Text table header
    _write_text_header(ws, row=2, start_col=text_col_start)

    # Text rows
    for i, tb in enumerate(blocks):
        _write_text_row(ws, row=3 + i, start_col=text_col_start, tb=tb)


# ── Table helpers ─────────────────────────────────────────────────────────

_TEXT_HEADERS = ["#", "Text", "Confidence", "Status", "Source"]


def _write_text_header(ws, row: int, start_col: int) -> None:
    for i, h in enumerate(_TEXT_HEADERS):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center")


def _write_text_row(ws, row: int, start_col: int, tb: TextBlock) -> None:
    values = [
        tb.order_index,
        tb.text,
        f"{tb.confidence:.2f}",
        tb.status,
        tb.source,
    ]
    fill = _pick_fill(tb)
    for i, val in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=val)
        if fill:
            cell.fill = fill
        if i == 1:  # Text column
            cell.alignment = Alignment(wrap_text=True)


def _pick_fill(tb: TextBlock):
    if tb.status == TextStatus.UNKNOWN.value:
        return _FILL_UNKNOWN
    if tb.status == TextStatus.SKIPPED.value:
        return _FILL_SKIP
    if tb.review_required:
        return _FILL_REVIEW
    return None
